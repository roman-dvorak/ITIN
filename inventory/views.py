import csv
import io
import json

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from django.views.generic import DetailView, ListView, TemplateView

from .access import can_edit_asset, visible_assets_for_user, visible_locations_for_user
from .forms import (
    AssetEditForm,
    AssetOSFeaturesForm,
    GuestSelfRegistrationForm,
    PortCreateForm,
    PortEditForm,
    PortInterfaceCreateForm,
    PortInterfaceEditForm,
)
from .models import (
    Asset,
    AssetOS,
    GuestDevice,
    IPAddress,
    Network,
    NetworkInterface,
    OrganizationalGroup,
    OSFamily,
    Location,
    Port,
)

User = get_user_model()


ASSET_EXPORT_COLUMNS = (
    "name",
    "asset_type",
    "status",
    "owner",
    "groups",
    "asset_tag",
    "serial_number",
    "manufacturer",
    "model",
    "notes",
    "metadata",
)
ASSET_IMPORT_COLUMNS = ASSET_EXPORT_COLUMNS


def user_can_edit_any_asset(user) -> bool:
    return bool(user.is_superuser or user.asset_admin_groups.exists())


def user_has_admin_access(user) -> bool:
    return bool(user.is_authenticated and user.is_staff)


def visible_users_for_user(user):
    queryset = User.objects.filter(is_active=True)
    if user.is_superuser:
        return queryset
    if user.is_staff:
        managed_groups = user.asset_admin_groups.all()
        return queryset.filter(
            Q(pk=user.pk) | Q(asset_member_groups__in=managed_groups) | Q(asset_admin_groups__in=managed_groups)
        ).distinct()
    return queryset.filter(pk=user.pk)


def visible_groups_for_user(user):
    if user.is_superuser:
        return OrganizationalGroup.objects.all()
    return OrganizationalGroup.objects.filter(Q(members=user) | Q(admins=user)).distinct()


def with_asset_table_related(queryset):
    interface_queryset = (
        NetworkInterface.objects.filter(active=True)
        .order_by("identifier")
        .prefetch_related(
            Prefetch(
                "ip_addresses",
                queryset=IPAddress.objects.filter(active=True).select_related("network"),
            )
        )
    )
    return queryset.select_related("owner", "location", "location__parent").prefetch_related(
        "groups",
        "os_entries__family",
        Prefetch("interfaces", queryset=interface_queryset),
    )


def filter_asset_queryset(queryset, params):
    q = params.get("q", "").strip()
    if hasattr(params, "getlist"):
        statuses = [value.strip() for value in params.getlist("status") if value.strip()]
        owners = [value.strip() for value in params.getlist("owner") if value.strip()]
        groups = [value.strip() for value in params.getlist("group") if value.strip()]
        os_families = [value.strip() for value in params.getlist("os_family") if value.strip()]
    else:
        statuses = [params.get("status", "").strip()]
        owners = [params.get("owner", "").strip()]
        groups = [params.get("group", "").strip()]
        os_families = [params.get("os_family", "").strip()]

    valid_statuses = [value for value in statuses if value in {choice[0] for choice in Asset.Status.choices}]
    valid_owner_ids = [int(value) for value in owners if value.isdigit()]
    valid_group_ids = [int(value) for value in groups if value.isdigit()]
    valid_os_family_ids = [int(value) for value in os_families if value.isdigit()]
    if q:
        queryset = queryset.filter(
            Q(name__icontains=q) | Q(asset_tag__icontains=q) | Q(serial_number__icontains=q)
        )
    if valid_statuses:
        queryset = queryset.filter(status__in=valid_statuses)
    if valid_owner_ids:
        queryset = queryset.filter(owner_id__in=valid_owner_ids)
    if valid_group_ids:
        queryset = queryset.filter(groups__id__in=valid_group_ids)
    if valid_os_family_ids:
        queryset = queryset.filter(os_entries__family_id__in=valid_os_family_ids)
    return queryset.order_by("name").distinct()


def _stringify_interface(interface):
    ip_values = [ip.address for ip in interface.ip_addresses.all()]
    ip_list = ",".join(ip_values) if ip_values else "-"
    return f"{interface.identifier}|{interface.mac_address or '-'}|{ip_list}"


def serialize_asset_row(asset):
    interfaces = "; ".join(_stringify_interface(interface) for interface in asset.interfaces.all())
    return [
        asset.name,
        asset.asset_type,
        asset.status,
        asset.owner.email if asset.owner_id else "",
        "; ".join(group.name for group in asset.groups.all()),
        asset.asset_tag,
        asset.serial_number,
        asset.manufacturer,
        asset.model,
        asset.notes,
        json.dumps(asset.metadata or {}, ensure_ascii=True),
        interfaces,
    ]


def _parse_groups(raw_value: str):
    if not raw_value:
        return []
    return [item.strip() for item in raw_value.split(";") if item.strip()]


def _resolve_owner(owner_value: str):
    owner_value = (owner_value or "").strip()
    if not owner_value:
        return None
    user = User.objects.filter(email__iexact=owner_value, is_active=True).first()
    if user:
        return user
    if owner_value.isdigit():
        return User.objects.filter(pk=int(owner_value), is_active=True).first()
    return User.objects.filter(username=owner_value, is_active=True).first()


def _update_asset_from_row(*, asset: Asset, row: dict, user, creating: bool):
    group_names = _parse_groups(row.get("groups", ""))
    if group_names:
        groups = list(OrganizationalGroup.objects.filter(name__in=group_names).distinct())
        missing = sorted(set(group_names) - {group.name for group in groups})
        if missing:
            raise ValueError(f"Unknown groups: {', '.join(missing)}.")
        if not user.is_superuser:
            allowed_ids = set(user.asset_admin_groups.values_list("id", flat=True))
            if any(group.id not in allowed_ids for group in groups):
                raise ValueError("Group assignment is outside your managed groups.")
    else:
        groups = []

    owner = _resolve_owner(row.get("owner", "")) or user
    if not owner:
        raise ValueError("Owner is required.")

    asset_type = (row.get("asset_type") or Asset.AssetType.COMPUTER).strip().upper()
    allowed_asset_types = {choice[0] for choice in Asset.AssetType.choices}
    if asset_type not in allowed_asset_types:
        raise ValueError(f"Invalid asset_type '{asset_type}'.")

    status = (row.get("status") or Asset.Status.ACTIVE).strip().upper()
    allowed_statuses = {choice[0] for choice in Asset.Status.choices}
    if status not in allowed_statuses:
        raise ValueError(f"Invalid status '{status}'.")

    metadata_raw = (row.get("metadata") or "").strip()
    if metadata_raw:
        try:
            metadata = json.loads(metadata_raw)
        except json.JSONDecodeError as error:
            raise ValueError("Metadata must be valid JSON.") from error
        if not isinstance(metadata, dict):
            raise ValueError("Metadata must be a JSON object.")
    else:
        metadata = {}

    if creating and not user.is_superuser and not groups:
        raise ValueError("Group is required for new assets.")

    asset.asset_type = asset_type
    asset.status = status
    asset.owner = owner
    asset.asset_tag = (row.get("asset_tag") or "").strip()
    asset.serial_number = (row.get("serial_number") or "").strip()
    asset.manufacturer = (row.get("manufacturer") or "").strip()
    asset.model = (row.get("model") or "").strip()
    asset.notes = (row.get("notes") or "").strip()
    asset.metadata = metadata
    asset.full_clean()
    asset.save()

    if groups:
        asset.groups.set(groups)
    elif creating:
        asset.groups.clear()


class HomeView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        assets = visible_assets_for_user(user)
        groups = visible_groups_for_user(user)
        now = timezone.now()

        if user.is_superuser:
            guest_devices = GuestDevice.objects.all()
        else:
            guest_devices = GuestDevice.objects.filter(Q(sponsor=user) | Q(groups__in=groups)).distinct()
        status_items = (
            assets.values("status")
            .annotate(total=Count("id"))
            .order_by("status")
        )
        status_map = {item["status"]: item["total"] for item in status_items}
        status_cards = [
            {"label": label, "code": code, "count": status_map.get(code, 0)}
            for code, label in Asset.Status.choices
        ]
        computers = assets.filter(asset_type__in=[Asset.AssetType.COMPUTER, Asset.AssetType.NOTEBOOK])
        active_computers = computers.filter(status=Asset.Status.ACTIVE).count()
        inactive_computers = computers.exclude(status=Asset.Status.ACTIVE).count()
        assets_without_os = assets.filter(os_entries__isnull=True).distinct().count()
        assets_with_os = assets.count() - assets_without_os
        assets_without_mac = assets.exclude(interfaces__mac_address__isnull=False).distinct().count()
        active_interfaces = NetworkInterface.objects.filter(asset__in=assets, active=True).count()
        active_ports = Port.objects.filter(asset__in=assets, active=True).count()
        os_items = (
            computers.filter(os_entries__isnull=False)
            .values("os_entries__family__family", "os_entries__family__name", "os_entries__family__flavor")
            .annotate(total=Count("id"))
            .order_by("os_entries__family__family", "os_entries__family__name", "os_entries__family__flavor")
        )
        family_labels = dict(OSFamily.FamilyType.choices)
        os_distribution = [
            {
                "family": family_labels.get(
                    item["os_entries__family__family"],
                    item["os_entries__family__family"] or "Other",
                ),
                "name": item["os_entries__family__name"],
                "flavor": item["os_entries__family__flavor"],
                "label": (
                    f"{item['os_entries__family__name']} - {item['os_entries__family__flavor']}"
                    if item["os_entries__family__flavor"]
                    else item["os_entries__family__name"]
                ),
                "count": item["total"],
            }
            for item in os_items
        ]
        without_os = computers.filter(os_entries__isnull=True).distinct().count()
        if without_os:
            os_distribution.append({"family": "No OS", "name": "Without OS", "flavor": None, "label": "Without OS", "count": without_os})
        asset_type_distribution = [
            {"label": label, "count": assets.filter(asset_type=code).count()}
            for code, label in Asset.AssetType.choices
        ]
        groups_with_counts = list(
            groups.annotate(total=Count("assets", filter=Q(assets__in=assets), distinct=True))
            .filter(total__gt=0)
            .order_by("-total", "name")
        )
        top_groups = groups_with_counts[:10]
        other_groups_total = sum(group.total for group in groups_with_counts[10:])
        group_distribution = [{"label": group.name, "count": group.total} for group in top_groups]
        if other_groups_total:
            group_distribution.append({"label": "Other", "count": other_groups_total})

        context.update(
            {
                "total_assets": assets.count(),
                "total_computers": computers.count(),
                "total_groups": groups.count(),
                "total_networks": Network.objects.count(),
                "active_guests": guest_devices.filter(
                    enabled=True,
                    approval_status=GuestDevice.ApprovalStatus.APPROVED,
                    valid_from__lte=now,
                    valid_until__gte=now,
                ).count(),
                "assets_with_ip": assets.filter(interfaces__ip_addresses__active=True).distinct().count(),
                "assets_with_os": assets_with_os,
                "assets_without_os": assets_without_os,
                "assets_without_mac": assets_without_mac,
                "active_interfaces": active_interfaces,
                "active_ports": active_ports,
                "status_cards": status_cards,
                "os_distribution": os_distribution,
                "asset_type_distribution": asset_type_distribution,
                "group_distribution": group_distribution,
                "computer_activity_distribution": [
                    {"label": "Active", "count": active_computers},
                    {"label": "Inactive", "count": inactive_computers},
                ],
            }
        )
        return context


class AssetListView(LoginRequiredMixin, ListView):
    model = Asset
    template_name = "inventory/asset_list.html"
    context_object_name = "assets"
    paginate_by = 25
    page_size_options = (25, 50, 100, 200)

    def get_paginate_by(self, queryset):
        raw_value = (self.request.GET.get("per_page") or "").strip()
        if raw_value.isdigit():
            page_size = int(raw_value)
            if page_size in self.page_size_options:
                return page_size
        return self.paginate_by

    def get_queryset(self):
        queryset = with_asset_table_related(visible_assets_for_user(self.request.user))
        return filter_asset_queryset(queryset, self.request.GET)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_queryset = self.object_list
        assets_for_table = list(current_queryset)
        for asset in assets_for_table:
            mac_addresses = [interface.mac_address for interface in asset.interfaces.all() if interface.mac_address]
            asset.mac_preview = mac_addresses[:2]
            asset.mac_extra_count = max(len(mac_addresses) - 2, 0)
            os_labels = []
            for os_entry in asset.os_entries.all():
                label = os_entry.family.name_flavor
                if os_entry.version:
                    label = f"{label} {os_entry.version}"
                os_labels.append(label)
            asset.os_display = ", ".join(os_labels)

        asset_ids = [asset.id for asset in assets_for_table]
        owners = User.objects.filter(owned_assets__id__in=asset_ids).distinct().order_by("email")
        groups = OrganizationalGroup.objects.filter(assets__id__in=asset_ids).distinct().order_by("name")
        os_families = (
            OSFamily.objects.filter(assetos__asset__id__in=asset_ids)
            .distinct()
            .order_by("family", "name", "flavor", "id")
        )
        params = self.request.GET.copy()
        params.pop("page", None)
        params_without_per_page = params.copy()
        params_without_per_page.pop("per_page", None)
        context["assets"] = assets_for_table
        context["total_filtered"] = context["paginator"].count if context.get("paginator") else len(assets_for_table)
        context["query"] = self.request.GET.get("q", "").strip()
        context["per_page"] = self.get_paginate_by(self.object_list)
        context["per_page_options"] = self.page_size_options
        context["status_filters"] = [value.strip() for value in self.request.GET.getlist("status") if value.strip()]
        context["owner_filters"] = [value.strip() for value in self.request.GET.getlist("owner") if value.strip()]
        context["group_filters"] = [value.strip() for value in self.request.GET.getlist("group") if value.strip()]
        context["os_family_filters"] = [value.strip() for value in self.request.GET.getlist("os_family") if value.strip()]
        context["status_choices"] = Asset.Status.choices
        context["owner_choices"] = owners
        context["group_choices"] = groups
        context["os_family_choices"] = os_families
        context["page_query"] = params.urlencode()
        context["page_query_without_per_page"] = params_without_per_page.urlencode()
        context["can_edit"] = user_can_edit_any_asset(self.request.user)
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        context["can_create"] = user_can_edit_any_asset(self.request.user)
        return context


class AssetCreateView(LoginRequiredMixin, View):
    template_name = "inventory/asset_create.html"

    def dispatch(self, request, *args, **kwargs):
        if not user_can_edit_any_asset(request.user):
            raise PermissionDenied("Missing permission to create assets.")
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        initial = {
            "owner": request.user.pk,
            "groups": list(request.user.asset_admin_groups.values_list("pk", flat=True)),
        }
        form = AssetEditForm(user=request.user, prefix="asset", initial=initial)
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = AssetEditForm(request.POST, user=request.user, prefix="asset")
        if form.is_valid():
            asset = form.save()
            messages.success(request, f"Asset {asset.name} created.")
            return redirect("inventory:asset-edit", pk=asset.pk)
        return render(request, self.template_name, {"form": form}, status=400)


class AssetExportView(LoginRequiredMixin, View):
    def get(self, request):
        export_format = request.GET.get("format", "csv").lower()
        queryset = filter_asset_queryset(
            with_asset_table_related(visible_assets_for_user(request.user)),
            request.GET,
        )
        if export_format == "xlsx":
            try:
                from openpyxl import Workbook
            except ImportError:
                messages.error(request, "XLSX export requires openpyxl.")
                return redirect("inventory:asset-list")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Assets"
            sheet.append(list(ASSET_EXPORT_COLUMNS) + ["interfaces"])
            for asset in queryset:
                sheet.append(serialize_asset_row(asset))
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="assets-export.xlsx"'
            return response

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="assets-export.csv"'
        writer = csv.writer(response)
        writer.writerow([*ASSET_EXPORT_COLUMNS, "interfaces"])
        for asset in queryset:
            writer.writerow(serialize_asset_row(asset))
        return response


class AssetImportView(LoginRequiredMixin, View):
    def post(self, request):
        if not user_can_edit_any_asset(request.user):
            raise PermissionDenied("Missing permission to import assets.")

        upload = request.FILES.get("file")
        if upload is None:
            messages.error(request, "Upload a CSV or XLSX file.")
            return redirect("inventory:asset-list")

        filename = upload.name.lower()
        try:
            if filename.endswith(".xlsx"):
                rows = self._rows_from_xlsx(upload)
            elif filename.endswith(".csv"):
                rows = self._rows_from_csv(upload)
            else:
                messages.error(request, "Unsupported format. Use .csv or .xlsx.")
                return redirect("inventory:asset-list")
        except RuntimeError as error:
            messages.error(request, str(error))
            return redirect("inventory:asset-list")
        except UnicodeDecodeError:
            messages.error(request, "CSV file must be UTF-8 encoded.")
            return redirect("inventory:asset-list")
        except Exception as error:  # noqa: BLE001
            messages.error(request, f"Import failed: {error}")
            return redirect("inventory:asset-list")

        if not rows:
            messages.error(request, "Import file is empty.")
            return redirect("inventory:asset-list")

        if "name" not in rows[0]:
            messages.error(request, "Missing required column: name.")
            return redirect("inventory:asset-list")

        created_count = 0
        updated_count = 0
        errors = []

        for index, row in enumerate(rows, start=2):
            name = (row.get("name") or "").strip()
            if not name:
                errors.append(f"Row {index}: name is required.")
                continue

            asset = Asset.objects.filter(name=name).first()
            creating = asset is None
            if creating:
                asset = Asset(name=name, owner=request.user)
            elif not can_edit_asset(request.user, asset):
                errors.append(f"Row {index}: no edit permission for asset '{name}'.")
                continue

            try:
                _update_asset_from_row(asset=asset, row=row, user=request.user, creating=creating)
            except ValueError as error:
                errors.append(f"Row {index}: {error}")
                continue
            if creating:
                created_count += 1
            else:
                updated_count += 1

        if created_count or updated_count:
            messages.success(request, f"Import finished: created {created_count}, updated {updated_count}.")
        if errors:
            for error in errors[:10]:
                messages.error(request, error)
            if len(errors) > 10:
                messages.error(request, f"{len(errors) - 10} more rows failed.")
        return redirect("inventory:asset-list")

    @staticmethod
    def _rows_from_csv(upload):
        decoded = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        return [dict(row) for row in reader]

    @staticmethod
    def _rows_from_xlsx(upload):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError("XLSX import requires openpyxl.") from error

        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(value).strip() if value is not None else "" for value in rows[0]]
        parsed = []
        for values in rows[1:]:
            row_data = {}
            for idx, key in enumerate(header):
                if not key:
                    continue
                value = values[idx] if idx < len(values) else None
                row_data[key] = "" if value is None else str(value)
            parsed.append(row_data)
        workbook.close()
        return parsed


class AssetImportTemplateView(LoginRequiredMixin, View):
    def get(self, _request):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="assets-import-template.csv"'
        writer = csv.writer(response)
        writer.writerow(ASSET_IMPORT_COLUMNS)
        return response


class AssetDetailView(LoginRequiredMixin, DetailView):
    model = Asset
    template_name = "inventory/asset_detail.html"
    context_object_name = "asset"

    def get_queryset(self):
        return (
            visible_assets_for_user(self.request.user)
            .select_related("owner", "location", "location__parent")
            .prefetch_related(
                "groups",
                "os_entries__family",
                "interfaces__ip_addresses__network",
                "ports__port_interfaces__ip_addresses__network",
            )
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        asset = context["asset"]
        context["can_edit"] = can_edit_asset(self.request.user, asset)
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        context["ports_tree"], context["unassigned_interfaces"] = build_ports_tree(asset)
        return context


class AssetOverviewView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/asset_overview.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["can_edit"] = user_can_edit_any_asset(self.request.user)
        return context


class OSCatalogView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/os_catalog.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        visible_assets = visible_assets_for_user(self.request.user)
        os_items = (
            OSFamily.objects.annotate(
                total_assets=Count(
                    "assetos__asset",
                    filter=Q(assetos__asset__in=visible_assets),
                    distinct=True,
                )
            )
            .order_by("family", "name", "flavor", "id")
        )
        context["os_items"] = os_items
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        return context


def build_location_tree(location_queryset):
    locations = list(location_queryset.select_related("parent").prefetch_related("groups"))
    nodes = {
        location.id: {
            "location": location,
            "children": [],
        }
        for location in locations
    }

    root_nodes = []
    for location in locations:
        node = nodes[location.id]
        if location.parent_id and location.parent_id in nodes:
            nodes[location.parent_id]["children"].append(node)
        else:
            root_nodes.append(node)

    def sort_node(node):
        node["children"].sort(key=lambda item: (item["location"].name.lower(), item["location"].id))
        for child in node["children"]:
            sort_node(child)

    root_nodes.sort(key=lambda item: (item["location"].name.lower(), item["location"].id))
    for root in root_nodes:
        sort_node(root)
    return root_nodes


class LocationTreeView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/location_tree.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        visible_locations = visible_locations_for_user(self.request.user)
        location_nodes = build_location_tree(visible_locations)
        location_ids = list(visible_locations.values_list("id", flat=True))
        assets_in_locations = (
            visible_assets_for_user(self.request.user)
            .filter(location_id__in=location_ids)
            .values("location_id")
            .annotate(total=Count("id"))
        )
        asset_counts = {item["location_id"]: item["total"] for item in assets_in_locations}

        def attach_counts(node):
            location = node["location"]
            node["asset_count"] = asset_counts.get(location.id, 0)
            for child in node["children"]:
                attach_counts(child)

        for root in location_nodes:
            attach_counts(root)

        context["location_nodes"] = location_nodes
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        return context


class LocationDetailView(LoginRequiredMixin, DetailView):
    model = Location
    template_name = "inventory/location_detail.html"
    context_object_name = "location"
    query_pk_and_slug = True

    def get_queryset(self):
        return visible_locations_for_user(self.request.user).prefetch_related("groups")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        location = context["location"]
        visible_locations = visible_locations_for_user(self.request.user)
        context["visible_children"] = visible_locations.filter(parent=location).prefetch_related("groups")
        location_assets = list(
            with_asset_table_related(
                visible_assets_for_user(self.request.user).filter(location=location)
            )
        )
        for asset in location_assets:
            mac_addresses = [interface.mac_address for interface in asset.interfaces.all() if interface.mac_address]
            asset.mac_preview = mac_addresses[:2]
            asset.mac_extra_count = max(len(mac_addresses) - 2, 0)
            os_labels = []
            for os_entry in asset.os_entries.all():
                label = os_entry.family.name_flavor
                if os_entry.version:
                    label = f"{label} {os_entry.version}"
                os_labels.append(label)
            asset.os_display = ", ".join(os_labels)
        context["location_assets"] = location_assets
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        return context


class LocationDetailLegacyRedirectView(LoginRequiredMixin, View):
    def get(self, request, pk):
        location = get_object_or_404(visible_locations_for_user(request.user), pk=pk)
        return redirect(location.get_absolute_url())


def get_editable_asset_or_403(user, pk):
    asset = get_object_or_404(
        visible_assets_for_user(user)
        .select_related("owner", "location", "location__parent")
        .prefetch_related("os_entries__family"),
        pk=pk,
    )
    if not can_edit_asset(user, asset):
        raise PermissionDenied("Missing edit permission for this asset.")
    return asset


def build_ports_tree(asset: Asset):
    ports_tree = []
    assigned_interface_ids = set()
    for port in asset.ports.filter(active=True).order_by("name"):
        interfaces = list(port.port_interfaces.filter(active=True).order_by("identifier"))
        for interface in interfaces:
            assigned_interface_ids.add(interface.id)
        ports_tree.append({"port": port, "interfaces": interfaces})
    unassigned_interfaces = [
        interface
        for interface in asset.interfaces.filter(active=True).order_by("identifier")
        if interface.id not in assigned_interface_ids
    ]
    return ports_tree, unassigned_interfaces


class AssetEditView(LoginRequiredMixin, View):
    template_name = "inventory/asset_edit.html"

    def get(self, request, pk):
        asset = get_editable_asset_or_403(request.user, pk)
        return render(request, self.template_name, self.get_context(asset))

    def post(self, request, pk):
        asset = get_editable_asset_or_403(request.user, pk)
        asset_form = AssetEditForm(request.POST, instance=asset, user=request.user, prefix="asset")
        if asset_form.is_valid():
            asset_form.save()
            messages.success(request, "Asset updated.")
            return redirect("inventory:asset-detail", pk=asset.pk)
        context = self.get_context(asset, asset_form=asset_form)
        return render(request, self.template_name, context, status=400)

    def get_context(self, asset, *, asset_form=None, os_form=None, port_form=None, os_row_forms=None):
        ports_tree, unassigned_interfaces = build_ports_tree(asset)
        for item in ports_tree:
            item["interface_form"] = PortInterfaceCreateForm(asset=asset, prefix=f"iface-{item['port'].id}")
            item["port_edit_form"] = PortEditForm(instance=item["port"], prefix=f"port-row-{item['port'].id}")
            item["interface_rows"] = [
                {
                    "interface": interface,
                    "form": PortInterfaceEditForm(
                        instance=interface,
                        asset=asset,
                        prefix=f"iface-row-{interface.id}",
                    ),
                }
                for interface in item["interfaces"]
            ]
        if os_row_forms is None:
            os_rows = list(asset.os_entries.select_related("family").order_by("-id"))
            os_row_forms = [
                {"record": row, "form": AssetOSFeaturesForm(instance=row, prefix=f"os-row-{row.id}")}
                for row in os_rows
            ]
        return {
            "asset": asset,
            "asset_form": asset_form or AssetEditForm(instance=asset, user=self.request.user, prefix="asset"),
            "os_form": os_form or AssetOSFeaturesForm(prefix="os-new"),
            "os_row_forms": os_row_forms,
            "port_form": port_form or PortCreateForm(prefix="port"),
            "ports_tree": ports_tree,
            "unassigned_interfaces": unassigned_interfaces,
        }


class AssetOSCreateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        asset = get_editable_asset_or_403(request.user, pk)
        form = AssetOSFeaturesForm(request.POST, prefix="os-new")
        if form.is_valid():
            if form.cleaned_data.get("family") is None:
                messages.error(request, "OS is required.")
            else:
                os_record = form.save(commit=False)
                os_record.asset = asset
                os_record.full_clean()
                os_record.save()
                messages.success(request, "OS entry added.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"OS {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class AssetOSUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk, os_id):
        asset = get_editable_asset_or_403(request.user, pk)
        os_record = get_object_or_404(AssetOS.objects.filter(asset=asset), pk=os_id)
        action = request.POST.get("action", "save")
        if action == "delete":
            os_record.delete()
            messages.success(request, "OS entry removed.")
            return redirect("inventory:asset-edit", pk=asset.pk)

        form = AssetOSFeaturesForm(request.POST, instance=os_record, prefix=f"os-row-{os_record.id}")
        if form.is_valid():
            if form.cleaned_data.get("family") is None:
                messages.error(request, "OS is required.")
            else:
                form.save()
                messages.success(request, "OS entry updated.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"OS {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class AssetPortCreateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        asset = get_editable_asset_or_403(request.user, pk)
        form = PortCreateForm(request.POST, prefix="port")
        if form.is_valid():
            port = form.save(commit=False)
            port.asset = asset
            port.active = True
            port.full_clean()
            port.save()
            messages.success(request, f"Port {port.name} created.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"Port {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class AssetPortInterfaceCreateView(LoginRequiredMixin, View):
    def post(self, request, pk, port_id):
        asset = get_editable_asset_or_403(request.user, pk)
        port = get_object_or_404(Port.objects.filter(asset=asset), pk=port_id)
        prefix = f"iface-{port.id}"
        identifier = request.POST.get(f"{prefix}-identifier", "").strip()
        existing_inactive = None
        if identifier:
            existing_inactive = asset.interfaces.filter(identifier=identifier, active=False).first()
        form = PortInterfaceCreateForm(
            request.POST,
            instance=existing_inactive,
            asset=asset,
            prefix=prefix,
        )
        if form.is_valid():
            interface = form.save(asset=asset, port=port)
            messages.success(request, f"Interface {interface.identifier} added under port {port.name}.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"Interface {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class AssetPortUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk, port_id):
        asset = get_editable_asset_or_403(request.user, pk)
        port = get_object_or_404(Port.objects.filter(asset=asset), pk=port_id)
        action = request.POST.get("action", "save")

        if action == "deactivate":
            port.active = False
            for interface in port.port_interfaces.filter(active=True):
                interface.active = False
                interface.save(update_fields=["active", "updated_at"])
                interface.ip_addresses.filter(active=True).update(active=False)
            port.save(update_fields=["active", "updated_at"])
            messages.success(request, f"Port {port.name} deactivated.")
            return redirect("inventory:asset-edit", pk=asset.pk)

        form = PortEditForm(request.POST, instance=port, prefix=f"port-row-{port.id}")
        if form.is_valid():
            updated_port = form.save(commit=False)
            updated_port.asset = asset
            updated_port.full_clean()
            updated_port.save()
            messages.success(request, f"Port {updated_port.name} updated.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"Port {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class AssetPortInterfaceUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk, port_id, interface_id):
        asset = get_editable_asset_or_403(request.user, pk)
        port = get_object_or_404(Port.objects.filter(asset=asset), pk=port_id)
        interface = get_object_or_404(asset.interfaces.filter(port=port), pk=interface_id)
        action = request.POST.get("action", "save")

        if action == "deactivate":
            interface.active = False
            interface.save(update_fields=["active", "updated_at"])
            interface.ip_addresses.filter(active=True).update(active=False)
            messages.success(request, f"Interface {interface.identifier} deactivated.")
            return redirect("inventory:asset-edit", pk=asset.pk)

        form = PortInterfaceEditForm(
            request.POST,
            instance=interface,
            asset=asset,
            prefix=f"iface-row-{interface.id}",
        )
        if form.is_valid():
            updated_interface = form.save()
            messages.success(request, f"Interface {updated_interface.identifier} updated.")
        else:
            for field, errors in form.errors.items():
                messages.error(request, f"Interface {field}: {' '.join(errors)}")
        return redirect("inventory:asset-edit", pk=asset.pk)


class GuestSelfRegisterView(View):
    template_name = "inventory/guest_register.html"

    def get(self, request):
        form = GuestSelfRegistrationForm(user=request.user)
        return render(request, self.template_name, {"form": form, "is_authenticated_user": request.user.is_authenticated})

    def post(self, request):
        form = GuestSelfRegistrationForm(request.POST, user=request.user)
        if form.is_valid():
            guest = form.save()
            messages.success(
                request,
                (
                    f"Registration submitted for {guest.mac_address}. "
                    f"Responsible person ({guest.sponsor.email}) must approve it."
                ),
            )
            return redirect("inventory:guest-register")
        return render(
            request,
            self.template_name,
            {"form": form, "is_authenticated_user": request.user.is_authenticated},
            status=400,
        )


class GuestApprovalListView(LoginRequiredMixin, TemplateView):
    template_name = "inventory/guest_approvals.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        user = self.request.user
        queryset = GuestDevice.objects.select_related("sponsor", "approved_by", "network").order_by("-created_at")
        if not user.is_superuser:
            queryset = queryset.filter(sponsor=user)

        context["pending_requests"] = queryset.filter(approval_status=GuestDevice.ApprovalStatus.PENDING)
        context["recent_requests"] = queryset[:100]
        context["active_count"] = queryset.filter(
            enabled=True,
            approval_status=GuestDevice.ApprovalStatus.APPROVED,
            valid_from__lte=now,
            valid_until__gte=now,
        ).count()
        context["now"] = now
        context["can_access_admin"] = user_has_admin_access(user)
        context["is_superuser"] = user.is_superuser
        return context


class GuestApproveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        guest = get_object_or_404(GuestDevice.objects.select_related("sponsor"), pk=pk)
        if not request.user.is_superuser and guest.sponsor_id != request.user.id:
            raise PermissionDenied("Missing permission to approve this guest request.")
        if guest.approval_status != GuestDevice.ApprovalStatus.PENDING:
            messages.error(request, "Only pending requests can be approved.")
            return redirect("inventory:guest-approvals")

        guest.approval_status = GuestDevice.ApprovalStatus.APPROVED
        guest.enabled = True
        guest.approved_by = request.user
        guest.approved_at = timezone.now()
        guest.rejected_reason = ""
        guest.save(
            update_fields=[
                "approval_status",
                "enabled",
                "approved_by",
                "approved_at",
                "rejected_reason",
                "updated_at",
            ]
        )
        messages.success(request, f"Guest device {guest.mac_address} approved.")
        return redirect("inventory:guest-approvals")


class GuestRejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        guest = get_object_or_404(GuestDevice.objects.select_related("sponsor"), pk=pk)
        if not request.user.is_superuser and guest.sponsor_id != request.user.id:
            raise PermissionDenied("Missing permission to reject this guest request.")
        if guest.approval_status != GuestDevice.ApprovalStatus.PENDING:
            messages.error(request, "Only pending requests can be rejected.")
            return redirect("inventory:guest-approvals")

        reason = request.POST.get("reason", "").strip()
        guest.approval_status = GuestDevice.ApprovalStatus.REJECTED
        guest.enabled = False
        guest.rejected_reason = reason
        guest.save(update_fields=["approval_status", "enabled", "rejected_reason", "updated_at"])
        messages.success(request, f"Guest device {guest.mac_address} rejected.")
        return redirect("inventory:guest-approvals")


class UserListView(LoginRequiredMixin, ListView):
    model = User
    template_name = "inventory/user_list.html"
    context_object_name = "users"
    paginate_by = 25

    def get_queryset(self):
        queryset = visible_users_for_user(self.request.user).prefetch_related(
            "asset_admin_groups", "asset_member_groups"
        ).order_by("email")
        q = self.request.GET.get("q", "").strip()
        if q and self.request.user.is_staff:
            queryset = queryset.filter(
                Q(email__icontains=q)
                | Q(first_name__icontains=q)
                | Q(last_name__icontains=q)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.request.GET.copy()
        params.pop("page", None)
        context["query"] = self.request.GET.get("q", "").strip()
        context["page_query"] = params.urlencode()
        context["only_profile"] = not self.request.user.is_staff
        context["profile_user"] = self.request.user
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        return context


class UserDetailView(LoginRequiredMixin, DetailView):
    model = User
    template_name = "inventory/user_detail.html"
    context_object_name = "profile_user"

    def get_queryset(self):
        return visible_users_for_user(self.request.user).prefetch_related(
            "asset_admin_groups",
            "asset_member_groups",
            "owned_assets",
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["can_access_admin"] = user_has_admin_access(self.request.user)
        context["owned_assets"] = with_asset_table_related(
            visible_assets_for_user(self.request.user).filter(owner=context["profile_user"])
        )
        return context
